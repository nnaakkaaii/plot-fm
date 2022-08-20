from pathlib import Path
from typing import Any, List, Optional

import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.chart import BarChart, Series, Reference

from pkg.env import OUTPUT_DIR


class WorkSheet:
    def __init__(self, ws: worksheet.Worksheet) -> None:
        self.__ws = ws
        self.__num_cols: Optional[int] = None
        self.__num_rows: Optional[int] = None

    def write(self,
              data: List[List[Any]],
              index: Optional[List[Any]] = None,
              header: Optional[List[Any]] = None,
              ) -> None:
        assert len(data) > 0 and len(data[0]) > 0
        if header is not None:
            for j, col in enumerate(header, start=2):
                self.__ws.cell(row=1, column=j, value=col)
        if index is not None:
            for i, ind in enumerate(index, start=2):
                self.__ws.cell(row=i, column=1, value=ind)
        for i, row in enumerate(data, start=2):
            for j, elem in enumerate(row, start=2):
                self.__ws.cell(row=i, column=j, value=elem)
        self.__num_rows = len(data)
        self.__num_cols = len(data[0])

    def set_col_width(self, width: int = 20) -> None:
        assert self.__num_cols is not None
        for i in range(1, 1 + self.__num_cols):
            self.__ws.column_dimensions[chr(65 + i)].width = width

    def cumulative_bar_chart(self,
                             title: str,
                             origin: str = 'D10',
                             x_axis_title: Optional[str] = None,
                             y_axis_title: Optional[str] = None,
                             ) -> None:
        assert self.__num_cols is not None
        assert self.__num_rows is not None
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.style = 12
        chart.overlap = 100
        chart.title = title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        data = Reference(self.__ws,
                         min_col=2,
                         min_row=1,
                         max_row=1 + self.__num_rows,
                         max_col=1 + self.__num_cols)
        cats = Reference(self.__ws,
                         min_col=1,
                         min_row=2,
                         max_row=1 + self.__num_rows)
        chart.add_data(data,
                       titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        self.__ws.add_chart(chart, origin)
        return


class WorkBook:
    def __init__(self) -> None:
        self.__wb = openpyxl.Workbook()
        self.__ws_ind = 0
        self.__path: Optional[Path] = None

    def save(self, name: str) -> None:
        assert isinstance(OUTPUT_DIR, str)
        p = Path(OUTPUT_DIR) / f'{name}.xlsx'
        self.__wb.save(p)
        self.__path = p

    def create_sheet(self, title: str) -> WorkSheet:
        if self.__ws_ind == 0:
            ws = self.__wb.active
            ws.title = title
        else:
            ws = self.__wb.create_sheet(index=self.__ws_ind, title=title)
        self.__ws_ind += 1
        return WorkSheet(ws)

    @property
    def path(self) -> Optional[Path]:
        return self.__path


def create() -> WorkBook:
    wb = WorkBook()
    return wb


if __name__ == '__main__':
    wb = create()
    ws1 = wb.create_sheet('このシートについて')
    ws1.write([['6532'], ['ベイカレント']],
              index=['企業ID', '企業名'])
    wb.save('test')
